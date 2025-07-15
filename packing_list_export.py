"""
packing_list_export.py
======================

Generate a “Packing List Print” Excel workbook from the consolidated
DataFrame in your Streamlit app.

Key features
------------
• Columns exactly in the order you specified  
• Dark‑blue subtotal row (only when a customer has > 1 item)  
• Grand totals calculated in the same columns  
• Sheet name: “Packing List Print”  
• Ready to be used with Streamlit’s download_button
"""

import io
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------- CONFIG -----------------------------------------------------------

HEADERS: List[str] = [
    "RECEIPT NO.", "MARK", "DESCRIPTION", "QTY",
    "MEAS. (CBM)", "WEIGHT(KG)", "WEIGHT RATE", "WEIGHT CBM",
    "CBM", "PER CHARGES", "TOTAL CHARGES", "CONTACT NUMBER",
    "业务员/ Supplier",
]

NUMERIC_TOTAL_COLS = {"QTY", "MEAS. (CBM)", "WEIGHT(KG)", "CBM", "TOTAL CHARGES"}

HEADER_FILL = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -----------------------------------------------------------------------------


def _safe_float(x):
    try:
        return float(x)
    except Exception:
        return 0.0


def export_custom_packing_list(df: pd.DataFrame) -> io.BytesIO:
    """
    Convert consolidated_df into an Excel file with per‑customer subtotals
    (blue) and an overall total row.

    Returns
    -------
    io.BytesIO
        Binary Excel content ready for Streamlit download.
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List Print"

    # ---- Header row ----
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL

    current_row = 2
    grand_totals = {k: 0.0 for k in NUMERIC_TOTAL_COLS}

    # Ensure all expected columns exist in df
    work_df = df.copy()
    for col in HEADERS:
        if col not in work_df.columns:
            work_df[col] = ""

    # ---- Process each customer (MARK) ----
    for customer, group in work_df.groupby("MARK", sort=False):
        item_rows = []

        # Split multi‑line fields into individual item rows
        for idx, g in group.iterrows():
            num_items = str(g["RECEIPT NO."]).count("\n") + 1
            receipts = str(g["RECEIPT NO."]).split("\n")
            descs = str(g["DESCRIPTION"]).split("\n")
            qtys = str(g["QTY"]).split("\n")
            meas_cbms = str(g["MEAS. (CBM)"] or g["CBM"]).split("\n")
            weights = str(g["WEIGHT(KG)"]).split("\n")

            weight_rate = _safe_float(g.get("WEIGHT RATE", 0))

            for i in range(num_items):
                item_rows.append({
                    "RECEIPT NO.": receipts[i] if i < len(receipts) else "",
                    "MARK": customer if i == 0 else "",
                    "DESCRIPTION": descs[i] if i < len(descs) else "",
                    "QTY": qtys[i] if i < len(qtys) else "",
                    "MEAS. (CBM)": meas_cbms[i] if i < len(meas_cbms) else "",
                    "WEIGHT(KG)": weights[i] if i < len(weights) else "",
                    "WEIGHT RATE": f"{weight_rate:.2f}" if i == 0 else "",
                    "WEIGHT CBM": f"{_safe_float(weights[i]) / weight_rate:.3f}" if weight_rate else "",
                    "CBM": g["CBM"].split("\n")[i] if "\n" in str(g["CBM"]) else g["CBM"],
                    "PER CHARGES": g["PER CHARGES"] if i == 0 else "",
                    "TOTAL CHARGES": g["TOTAL CHARGES"] if i == 0 else "",
                    "CONTACT NUMBER": g["CONTACT NUMBER"] if i == 0 else "",
                    "业务员/ Supplier": g["业务员/ Supplier"] if i == 0 else "",})

        # Write item rows
        for item in item_rows:
            ws.append([item[col] for col in HEADERS])
            for col_name in NUMERIC_TOTAL_COLS:
                grand_totals[col_name] += _safe_float(item[col_name])
            current_row += 1

        # Subtotal row (only if >1 item)
        if len(item_rows) > 1:
            subtotal_vals = {k: 0.0 for k in NUMERIC_TOTAL_COLS}
            for item in item_rows:
                for k in NUMERIC_TOTAL_COLS:
                    subtotal_vals[k] += _safe_float(item[k])

            ws.append([
                f"{customer} TOTAL", "", "",                           # first 3 cols
                f"{subtotal_vals['QTY']:.2f}",
                f"{subtotal_vals['MEAS. (CBM)']:.3f}",
                f"{subtotal_vals['WEIGHT(KG)']:.2f}",
                "", "", f"{subtotal_vals['CBM']:.3f}", "",
                f"{subtotal_vals['TOTAL CHARGES']:.2f}", "", ""
            ])
            # Color subtotal row
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = SUBTOTAL_FILL
                cell.font = Font(bold=True)
            current_row += 1

    # ---- Grand total row ----
    ws.append([
        "GRAND TOTAL", "", "", f"{grand_totals['QTY']:.2f}",
        f"{grand_totals['MEAS. (CBM)']:.3f}",
        f"{grand_totals['WEIGHT(KG)']:.2f}", "", "",
        f"{grand_totals['CBM']:.3f}", "",
        f"{grand_totals['TOTAL CHARGES']:.2f}", "", ""
    ])
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = Font(bold=True)

    # ---- Auto‑size columns ----
    for col in ws.columns:
        first_cell = next((c for c in col if c.value not in (None, "")), None)
        if first_cell:
            width = min(25, len(str(first_cell.value)) + 2)
            ws.column_dimensions[get_column_letter(first_cell.column)].width = max(10, width)

    wb.save(buffer)
    buffer.seek(0)
    return buffer
