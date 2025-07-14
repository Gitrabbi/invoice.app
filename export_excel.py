# export_excel.py

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from io import BytesIO

SEA_BLUE = PatternFill(start_color="B7D6F4", end_color="B7D6F4", fill_type="solid")

EXPORT_COLUMNS = [
    "RECEIPT NO.", "MARK", "DESCRIPTION", "QTY",
    "MEAS.(CBM)", "WEIGHT(KG)", "WEIGHT RATE", "WEIGHT CBM",
    "CBM", "PER CHARGES", "TOTAL CHARGES", "CONTACT NUMBER",
    "业务员/ Supplier"
]

NUMERIC_TOTAL_COLS = ["QTY", "MEAS.(CBM)", "WEIGHT(KG)", "CBM", "TOTAL CHARGES"]

def export_to_excel_with_totals(df: pd.DataFrame) -> BytesIO:
    """
    Create an Excel file from df with:
    - Data written in a fixed column order
    - A blue total row after each customer group (if >1 row)
    - A grand total row at the bottom
    Sheet is named 'Packing List Print'.
    Returns a BytesIO stream ready for download.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List Print"

    # --- Write header ---
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    current_row = 2
    grand_totals = {c: 0 for c in NUMERIC_TOTAL_COLS}

    # fill missing columns
    work_df = df.copy()
    for col in EXPORT_COLUMNS:
        if col not in work_df.columns:
            work_df[col] = ""

    # group by customer
    grouped = work_df.groupby("MARK", sort=False)

    for customer, group in grouped:
        for _, row in group.iterrows():
            for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
                ws.cell(row=current_row, column=col_idx, value=row[col_name])
            # accumulate grand totals
            for col in NUMERIC_TOTAL_COLS:
                try:
                    grand_totals[col] += float(row[col])
                except ValueError:
                    pass
            current_row += 1

        # Add customer total row if more than one item
        if len(group) > 1:
            ws.cell(row=current_row, column=2, value=f"{customer} TOTAL")
            for col_name in NUMERIC_TOTAL_COLS:
                col_idx = EXPORT_COLUMNS.index(col_name) + 1
                ws.cell(row=current_row, column=col_idx, value=group[col_name].sum())
            # Sea blue fill
            for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
                ws.cell(row=current_row, column=col_idx).fill = SEA_BLUE
            current_row += 1

    # --- Final grand total row ---
    ws.cell(row=current_row, column=2, value="GRAND TOTAL")
    for col_name in NUMERIC_TOTAL_COLS:
        col_idx = EXPORT_COLUMNS.index(col_name) + 1
        ws.cell(row=current_row, column=col_idx, value=grand_totals[col_name])
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True)

    # Auto-size columns
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].bestFit = True
        ws.column_dimensions[col_letter].auto_size = True

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
